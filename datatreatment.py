from openpyxl import Workbook
import openpyxl


def prepost(path):
    file_path = "MondayBot\PostDB.xlsx"
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=3):
        cell_a = row[0].value  # Column A
        cell_b = row[1].value  # Column B
        cell_c = row[2].value  # Column C
        pos_a = row[0].coordinate
        pos_b = row[1].coordinate
        pos_c = row[2].coordinate
        

        if cell_b == "Ready" and cell_c != None:
            sheet[str(pos_b)] = "Posted"
            wb.save("data.xlsx") 
            wb.close()
            return cell_c
        
        
print(prepost("MondayBot\PostDB.xlsx"))
        
    
    
